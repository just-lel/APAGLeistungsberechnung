# app.py
from __future__ import annotations

import io
import re
from copy import copy as _copy
from dataclasses import dataclass
from typing import Dict, Optional, Tuple, List

import pandas as pd
import streamlit as st
from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet


# -----------------------------
# Defaults (kannst du in der UI ändern)
# -----------------------------
DEFAULT_START_ROW_BY_VERTEILUNG = {
    "HV": 25,
    "1": 13,
    "2": 13,
    "3": 13,
    "4": 13,
    "5": 13,
}
DEFAULT_SHEETNAME_BY_VERTEILUNG = {k: k for k in DEFAULT_START_ROW_BY_VERTEILUNG.keys()}

# Excel Zielspalten gemäß deinem Flussdiagramm:
# A: Typ + Verwendung, B: Anzahl, C: Leistung, E: Gleichzeitigkeit
EXCEL_COL_A = 1
EXCEL_COL_B = 2
EXCEL_COL_C = 3
EXCEL_COL_E = 5


# -----------------------------
# Spalten-Mapping (CSV -> Standard)
# -----------------------------
COLUMN_ALIASES: Dict[str, Tuple[str, ...]] = {
    "Anzahl": ("Anzahl", "anzahl", "\ufeffAnzahl", "quantity"),
    "Typ": ("Typ", "typ", "Type"),
    "Leistung": ("Leistung", "leistung", "Power", "kW", "kw"),
    "Gleichzeitigkeit": ("Gleichzeitigkeit", "gleichzeitigkeit", "Gleichzeitigkeitsfaktor", "simultaneity", "GF"),
    "Ebene": ("Ebene", "ebene", "Level", "Geschoss", "geschoss"),
    "Verteilung": ("Verteilung", "verteilung", "Distribution", "distribution"),
    "Verwendung": ("Verwendung", "verwendung", "BSAG_Verwendung", "Nutzung", "nutzung"),
}


def _to_str(x) -> str:
    if pd.isna(x):
        return ""
    return str(x).strip()


def _safe_float(x, default: float = 0.0) -> float:
    if pd.isna(x):
        return default
    if isinstance(x, (int, float)):
        return float(x)
    s = str(x).strip().replace(",", ".")
    try:
        return float(s)
    except ValueError:
        return default


def _parse_verteilung(v) -> str:
    s = _to_str(v)
    return s.upper() if s.upper() == "HV" else s


def normalize_columns(df: pd.DataFrame) -> pd.DataFrame:
    col_map = {}
    lower_cols = {c.lower(): c for c in df.columns}

    def find_col(target: str) -> Optional[str]:
        for alias in COLUMN_ALIASES[target]:
            key = alias.lower()
            if key in lower_cols:
                return lower_cols[key]
        return None

    for target in COLUMN_ALIASES.keys():
        src = find_col(target)
        if src is not None:
            col_map[src] = target

    df = df.rename(columns=col_map)

    required = ["Anzahl", "Typ", "Leistung", "Gleichzeitigkeit", "Ebene", "Verteilung"]
    missing = [c for c in required if c not in df.columns]
    if missing:
        raise ValueError(
            f"Fehlende Spalten in CSV: {missing}\n"
            f"Gefunden: {list(df.columns)}\n"
            f"Passe ggf. COLUMN_ALIASES an."
        )

    if "Verwendung" not in df.columns:
        df["Verwendung"] = ""

    return df


# -----------------------------
# Ebene-Logik
# -----------------------------
def suggest_level_short(level_raw: str) -> str:
    """
    Vorschlag für Kurzform (kann der User im UI überschreiben)
    """
    s = _to_str(level_raw).upper()

    if any(k in s for k in ["UG", "KG", "UNTERGESCHOSS", "KELLER"]):
        n = re.findall(r"(\d+)", s)
        return f"UG{n[0]}" if n else "UG"

    if any(k in s for k in ["EG", "ERDGESCHOSS"]):
        return "EG"

    if any(k in s for k in ["OG", "OBERGESCHOSS"]):
        n = re.findall(r"(\d+)", s)
        return f"OG{n[0]}" if n else "OG"

    s = re.sub(r"\s+", " ", s).strip()
    return s if s else "UNBEKANNT"


def level_sort_key(level_short: str) -> Tuple[int, int, str]:
    """
    Standard-Sortierung: UG (tiefer zuerst) -> EG -> OG (aufsteigend) -> Rest
    """
    s = _to_str(level_short).upper()

    m = re.match(r"UG(\d+)?$", s)
    if m:
        n = int(m.group(1)) if m.group(1) else 1
        return (-100, -n, s)

    if s == "EG":
        return (0, 0, s)

    m = re.match(r"OG(\d+)?$", s)
    if m:
        n = int(m.group(1)) if m.group(1) else 1
        return (100, n, s)

    return (1000, 0, s)


# -----------------------------
# Excel Writing Helpers
# -----------------------------
def copy_row_style(ws: Worksheet, src_row: int, dst_row: int, max_col: int = 30) -> None:
    for col in range(1, max_col + 1):
        src_cell = ws.cell(row=src_row, column=col)
        dst_cell = ws.cell(row=dst_row, column=col)
        if src_cell.has_style:
            dst_cell._style = _copy(src_cell._style)
        dst_cell.number_format = src_cell.number_format


def find_first_empty_row(ws: Worksheet, start_row: int, col: int = EXCEL_COL_A, scan_limit: int = 2000) -> int:
    r = start_row
    while r < start_row + scan_limit:
        if ws.cell(row=r, column=col).value in (None, ""):
            return r
        r += 1
    raise RuntimeError("Konnte keine freie Zeile finden (Scan-Limit erreicht).")


# -----------------------------
# Transformation
# -----------------------------
@dataclass(frozen=True)
class ExportConfig:
    sheetname_by_verteilung: Dict[str, str]
    startrow_by_verteilung: Dict[str, int]
    custom_level_order: Optional[List[str]]  # wenn gesetzt, nimmt diese Reihenfolge


def prepare_dataframe(
    df: pd.DataFrame,
    level_map: Dict[str, str],
    cfg: ExportConfig,
) -> pd.DataFrame:
    df = normalize_columns(df).copy()

    df["Verteilung"] = df["Verteilung"].apply(_parse_verteilung)
    df["Ebene"] = df["Ebene"].apply(_to_str)
    df["Ebene_short"] = df["Ebene"].map(lambda x: level_map.get(x, suggest_level_short(x)))

    df["Typ"] = df["Typ"].apply(_to_str)
    df["Verwendung"] = df["Verwendung"].apply(_to_str)
    df["Text_A"] = (df["Typ"] + " " + df["Verwendung"]).str.strip()

    df["Anzahl"] = df["Anzahl"].apply(_safe_float)
    df["Leistung"] = df["Leistung"].apply(_safe_float)
    df["Gleichzeitigkeit"] = df["Gleichzeitigkeit"].apply(_safe_float)


# Leistung ist Einzelleistung -> NICHT summieren!
# Wir nehmen den ersten Wert innerhalb der Gruppe.
# (Optional: nimm "median" oder "max", wenn das bei dir besser passt.)
    grouped = (
        df.groupby(["Verteilung", "Ebene_short", "Text_A"], as_index=False)
        .agg({
            "Anzahl": "sum",
            "Leistung": "first",          # <- Einzelleistung beibehalten
            "Gleichzeitigkeit": "first",   # bleibt wie gewählt (mean/max/first)
        })
    )

    # Sortierung
    if cfg.custom_level_order:
        order_index = {lvl: i for i, lvl in enumerate(cfg.custom_level_order)}
        grouped["__lvl_order__"] = grouped["Ebene_short"].map(lambda x: order_index.get(x, 10_000))
        grouped = grouped.sort_values(["Verteilung", "__lvl_order__", "Text_A"]).drop(columns="__lvl_order__")
    else:
        grouped["__sortkey__"] = grouped["Ebene_short"].apply(level_sort_key)
        grouped = grouped.sort_values(["Verteilung", "__sortkey__", "Text_A"]).drop(columns="__sortkey__")

    return grouped


def write_to_template_bytes(
    prepared: pd.DataFrame,
    template_bytes: bytes,
    cfg: ExportConfig,
) -> bytes:
    wb = load_workbook(io.BytesIO(template_bytes))

    for verteilung, sub in prepared.groupby("Verteilung"):
        v = str(verteilung)

        if v not in cfg.sheetname_by_verteilung or v not in cfg.startrow_by_verteilung:
            continue

        sheet_name = cfg.sheetname_by_verteilung[v]
        start_row = cfg.startrow_by_verteilung[v]

        if sheet_name not in wb.sheetnames:
            continue

        ws = wb[sheet_name]
        template_style_row = start_row

        # Erste freie Zeile ab Start
        r = find_first_empty_row(ws, start_row, col=EXCEL_COL_A)

        # Wir iterieren Ebene-weise (blockweise)
        # Wichtig: sub ist bereits nach Ebene sortiert (über prepare_dataframe)
        current_level = None

        for _, row in sub.iterrows():
            level = row["Ebene_short"]

            # Wenn neue Ebene beginnt -> (optional) Leerzeile + Ebene als Header-Zeile
            if level != current_level:
                # <-- HIER: Leerzeile zwischen Ebenen (aber nicht vor der ersten Ebene)
                if current_level is not None:
                    r += 1

                current_level = level

                # Header-Zeile schreiben
                if r != template_style_row:
                    copy_row_style(ws, template_style_row, r, max_col=30)

                ws.cell(row=r, column=EXCEL_COL_A, value=str(level))
                ws.cell(row=r, column=EXCEL_COL_B, value=None)
                ws.cell(row=r, column=EXCEL_COL_C, value=None)
                ws.cell(row=r, column=EXCEL_COL_E, value=None)

                ws.cell(row=r, column=EXCEL_COL_A).font = ws.cell(
                    row=template_style_row, column=EXCEL_COL_A
                ).font.copy(bold=True)

                r += 1

            # Objekt-Zeile schreiben
            if r != template_style_row:
                copy_row_style(ws, template_style_row, r, max_col=30)

            ws.cell(row=r, column=EXCEL_COL_A, value=row["Text_A"])
            ws.cell(row=r, column=EXCEL_COL_B, value=float(row["Anzahl"]))
            ws.cell(row=r, column=EXCEL_COL_C, value=float(row["Leistung"]) / 1000.0)
            ws.cell(row=r, column=EXCEL_COL_E, value=float(row["Gleichzeitigkeit"]))
            r += 1

    out = io.BytesIO()
    wb.save(out)
    return out.getvalue()


# -----------------------------
# Streamlit UI
# -----------------------------
st.set_page_config(page_title="CSV → Excel Vorlage (Verteilungen)", layout="wide")

st.title("CSV → Excel Vorlage")
st.caption("Upload, Ebenen zuordnen, Export in Excel-Vorlage, Download.")

with st.sidebar:
    st.header("1) Dateien")
    csv_file = st.file_uploader("CSV hochladen", type=["csv"])
    template_file = st.file_uploader("Excel-Vorlage (.xlsx) hochladen", type=["xlsx"])

    st.header("3) Verteilungen → Blatt & Startzeile")
    st.write("Standard: HV ab Zeile 25, Verteilung 1..5 ab Zeile 13 (anpassbar).")

    # Editierbare Config
    if "sheet_cfg" not in st.session_state:
        st.session_state.sheet_cfg = DEFAULT_SHEETNAME_BY_VERTEILUNG.copy()
    if "start_cfg" not in st.session_state:
        st.session_state.start_cfg = DEFAULT_START_ROW_BY_VERTEILUNG.copy()

    # Konfig-Tabelle
    cfg_rows = []
    keys = sorted(set(st.session_state.sheet_cfg.keys()) | set(st.session_state.start_cfg.keys()))
    for k in keys:
        cfg_rows.append({"Verteilung": k, "Sheet": st.session_state.sheet_cfg.get(k, ""), "Startzeile": st.session_state.start_cfg.get(k, 13)})

    cfg_df = pd.DataFrame(cfg_rows)
    edited_cfg = st.data_editor(cfg_df, num_rows="dynamic", use_container_width=True)

    # zurückschreiben
    sheet_cfg = {}
    start_cfg = {}
    for _, r in edited_cfg.iterrows():
        v = _to_str(r.get("Verteilung"))
        if not v:
            continue
        sheet_cfg[v] = _to_str(r.get("Sheet")) or v
        try:
            start_cfg[v] = int(r.get("Startzeile"))
        except Exception:
            start_cfg[v] = 13

    st.session_state.sheet_cfg = sheet_cfg
    st.session_state.start_cfg = start_cfg

    st.header("4) Ebenen sortieren")
    use_custom_order = st.checkbox("Eigene Ebenen-Reihenfolge verwenden", value=False)
    custom_order_text = st.text_area(
        "Reihenfolge (eine Ebene pro Zeile, exakt wie Kurzform)",
        value="UG2\nUG1\nUG\nEG\nOG1\nOG2\nOG3",
        height=140,
        disabled=not use_custom_order,
    )

# Lade CSV -> DataFrame
df_raw = None
if csv_file is not None:
    try:
        df_raw = pd.read_csv(csv_file, sep=None, engine="python")
    except Exception as e:
        st.error(f"CSV konnte nicht gelesen werden: {e}")

if df_raw is None:
    st.info("Bitte CSV und Excel-Vorlage hochladen, dann Ebenen zuordnen und exportieren.")
    st.stop()

# Normalisieren (für Ebenen-Liste brauchen wir Ebene-Spalte)
try:
    df_norm = normalize_columns(df_raw.copy())
except Exception as e:
    st.error(str(e))
    st.stop()

# Ebenen auslesen
unique_levels = sorted({_to_str(x) for x in df_norm["Ebene"].dropna().unique() if _to_str(x)})

st.subheader("Ebenen-Zuordnung")
st.write("Ordne jede gefundene Ebene einer Kurzform zu (z. B. UG, EG, OG1). Vorschläge sind vorausgefüllt.")

# Session state für Ebenen-Mapping
if "level_map" not in st.session_state:
    st.session_state.level_map = {lvl: suggest_level_short(lvl) for lvl in unique_levels}

# UI: Tabelle editierbar
level_df = pd.DataFrame(
    [{"Ebene (CSV)": lvl, "Kurzform": st.session_state.level_map.get(lvl, suggest_level_short(lvl))} for lvl in unique_levels]
)
edited_level_df = st.data_editor(level_df, use_container_width=True, hide_index=True)

level_map = {}
for _, r in edited_level_df.iterrows():
    src = _to_str(r.get("Ebene (CSV)"))
    dst = _to_str(r.get("Kurzform"))
    if src:
        level_map[src] = dst if dst else suggest_level_short(src)

st.session_state.level_map = level_map

# Vorschau
col1, col2 = st.columns([1, 1])
with col1:
    st.subheader("CSV Vorschau")
    st.dataframe(df_norm.head(50), use_container_width=True)

with col2:
    st.subheader("Gefundene Verteilungen")
    vlist = sorted({_parse_verteilung(v) for v in df_norm["Verteilung"].dropna().unique()})
    st.write(vlist)

# Export-Config
custom_order = None
if use_custom_order:
    custom_order = [line.strip() for line in custom_order_text.splitlines() if line.strip()]

cfg = ExportConfig(
    sheetname_by_verteilung=st.session_state.sheet_cfg,
    startrow_by_verteilung=st.session_state.start_cfg,
    custom_level_order=custom_order,
)

# Button: Export
st.divider()
st.subheader("Export")

if template_file is None:
    st.warning("Bitte noch eine Excel-Vorlage (.xlsx) hochladen.")
    st.stop()

do_export = st.button("Excel erzeugen", type="primary")

if do_export:
    try:
        prepared = prepare_dataframe(df_norm, level_map=level_map, cfg=cfg)
        st.success(f"Aufbereitete Zeilen: {len(prepared)}")

        st.subheader("Aufbereitete Daten (Preview)")
        st.dataframe(prepared.head(200), use_container_width=True)

        out_bytes = write_to_template_bytes(prepared, template_bytes=template_file.getvalue(), cfg=cfg)

        st.download_button(
            label="⬇️ Excel herunterladen",
            data=out_bytes,
            file_name="export.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
    except Exception as e:
        st.error(f"Fehler beim Export: {e}")
